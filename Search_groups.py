import os
import openpyxl


def Search_groups():
    directory = 'Admin/Downloads'
    search_list = ['АМ-31,33', 'АМ-32,34', 'АМ-41', 'АМ-42,43', 'АС-11', 'АЭ-11', 'АЭ-21,22', 'АЭ-31,32', 'БУ-21,22',
                   'БУ-31,32', 'ИС-11', 'ИС-12', 'ИС-13', 'ИС-21,23', 'ИС-22,24', 'ИС-31,34',
                   'ИС-32,35', 'ИС-33,36', 'ИС-41,43', 'ИС-42,44', 'ЛЗ-11', 'ЛЗ-21,22', 'ЛЗ-31,32', 'ЛЗ-41', 'ЛП-11',
                   'МЭ-11', 'МЭ-21,22', 'МЭ-31,32', 'МЭ-41,42', 'ОЛ-11', 'ОЛ-21,22', 'ОЛ-31,32',
                   'ОП-11', 'ОП-21,22', 'ОП-31,32', 'ОП-41,44', 'РД-11', 'РЗ-11', 'ТД-11', 'ТД-21,22', 'ТД-31,32',
                   'ТД-41,42', 'ТП-11', 'ТП-21', 'ТП-31,32', 'ТП-41', 'ЭМ-11', 'ЭМ-12', 'ЭМ-21,22',
                   'ЭМ-31,32', 'ЭМ-41,43', 'ЭМ-42,44', 'ЭС-11', 'ЭС-21', 'ЭС-31,32', 'ЭС-41,42', 'ЮР-11,13', 'ЮР-12',
                   'ЮР-21,23', 'ЮР-22,24', 'ЮР-25', 'ЮР-31,33', 'ЮР-32,34', 'ЮР-35,36']

    values_dict = {}

    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            values = []
            filepath = os.path.join(directory, filename)
            wb = openpyxl.load_workbook(filepath)
            for sheet in wb.worksheets:
                for row in sheet.rows:
                    for cell in row:
                        if cell.value in search_list:
                            values.append((cell.value, cell.coordinate))
            values_dict[filename] = values

    # print(values_dict)

    # Получаем список файлов с расширением xlsx
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]

    # Создаем коллекцию значений для каждого файла
    for file in files:
        wb = openpyxl.load_workbook(os.path.join(directory, file))
        sheet = wb.active

        # Создаем словарь для текущего файла
        cell_values = {}

        # Перебираем все ячейки и собираем информацию
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    for group in search_list:
                        if group in value:
                            cell_values[group] = cell.coordinate

        # Создаем текстовый файл с названием текущего xlsx файла
        with open(os.path.join(directory, os.path.splitext(file)[0] + '.txt'), 'w') as f:
            for group, cell in cell_values.items():
                f.write(f"{group}: {cell}\n")
