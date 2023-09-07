import os
import openpyxl as ox
import numpy as np
import time
import json
import datetime
import re
from datetime import datetime, date, timedelta

def Delete_old_files():
    data_dir = 'Admin/Downloads'
    file_list = os.listdir(data_dir)
    date_pattern = r'\d{2}\.\d{2}'
    date_list = []
    for filename in file_list:
        # Проверяем, соответствует ли имя файла формату с датой
        date_match = re.search(date_pattern, filename)

        if date_match:
            # Извлекаем дату из имени файла
            date_str = date_match.group(0)
            # Проверяем, есть ли слово "НОВОЕ" в имени файла
            if ("НОВОЕ" or "новое") in filename:
                # Если есть, удаляем все старые файлы с этой датой
                for old_file in file_list:
                    if date_str in old_file and ("НОВОЕ" or "новое") not in old_file:
                        try:
                            print(old_file, "deleted")
                            os.remove(os.path.join(data_dir, old_file))
                        except Exception:
                            continue
        # print(date_list)
            # else:
            #     # Если слова "НОВОЕ" нет, удаляем все файлы с этой датой, кроме текущего
            #     for old_file in file_list:
            #         if date_str in old_file and filename != old_file:
            #             os.remove(os.path.join(data_dir, old_file))


def Reader_files():
    start_reader = time.time()
    all_data_collect = {}
    data_dir = 'Admin/Downloads'
    txt_files = [f for f in os.listdir(data_dir) if f.endswith('.txt')]
    xlsx_files = [v for v in os.listdir(data_dir) if v.endswith('.xlsx')]
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    dest_dir = 'Admin/Destination'
    fds = sorted(os.listdir('Admin/Destination/'))
    for file in fds:
        if file.endswith('.txt'):
            os.remove(dest_dir + '/' + file)

    for file in txt_files:
        data = {}
        with open(os.path.join(data_dir, file), 'r') as f:
            lines = f.readlines()
            for line in lines:
                line = line.strip().split(': ')
                data[line[0]] = line[1]
        print(file)

        for filename in xlsx_files:
            if filename.endswith('.xlsx') and (str(filename)[:-4] + "txt") == file:

                current_year = str(datetime.now().year)
                date_str = []
                date_start = file.find(" на ") + 4
                date_end = file.find(".2023")
                date_str.append(file[date_start:date_end])

                date_pattern = r'\d{2}\.\d{2}'
                date_match = re.search(date_pattern, file)
                if date_match:
                    # Извлекаем дату из имени файла
                    date_str = str(date_match.group(0)) + "." + current_year
                # if "НОВОЕ" or "новое" in str(filename):
                #     date_str = str(date_str).replace('[', '').replace(']', '').replace("'", "")[:-2] + current_year
                # else:
                #     date_str = str(date_str).replace('[', '').replace(']', '').replace("'", "")[:-2] + current_year
                all_data_collect[str(date_str)] = {}

                start = time.time()
                # print(filename)
                filepath = os.path.join(data_dir, filename)
                wb = ox.load_workbook(filename=filepath, read_only=True)
                ws = wb[wb.sheetnames[0]]
                row_max = ws.max_row
                column_max = ws.max_column
                print('In file: ', filename, '\n Columns:', column_max, '\n Rows: ', row_max)
                row_min = 1
                column_min = 1

                ########################################################################################################

                arr_A = []
                arr_indexes_A = []
                arr_indexes_without_A = []
                for i in range(row_max + 1):
                    word_cell_A = 'A' + str(int(i))
                    word_cell_without_A = int(i)
                    arr_indexes_A.append(word_cell_A)
                    arr_indexes_without_A.append(word_cell_without_A)
                    # print(word_cell_A)
                    data_from_cell_A = ws[word_cell_A].value
                    arr_A.append(data_from_cell_A)

                NoneType = type(None)
                index_to_remove = []
                for j in range(len(arr_A)):
                    if type(arr_A[j]) == NoneType:
                        index_to_remove.append(j)

                arr_A = np.delete(arr_A, index_to_remove)
                arr_indexes_A = np.delete(arr_indexes_A, index_to_remove)
                arr_indexes_without_A = np.delete(arr_indexes_without_A, index_to_remove)

                # Очистим массивы индексов и данных из ячеек от мусора
                next_clean_arr = []
                for i in range(len(arr_A)):
                    nums = '123456'
                    if str(arr_A[i]) not in nums:
                        next_clean_arr.append(i)
                arr_A = np.delete(arr_A, next_clean_arr)
                arr_indexes_A = np.delete(arr_indexes_A, next_clean_arr)
                arr_indexes_without_A = np.delete(arr_indexes_without_A, next_clean_arr)
                # Функция создания массивов индексов и данных из ячеек
                arr_test = arr_A
                arr_test_ind = arr_indexes_A
                arr_test_ind_without_A = arr_indexes_without_A

                result_arr = []
                result_arr_ind = []
                result_arr_ind_without_A = []

                temp = [arr_test[0]]
                temp_ind = [arr_test_ind[0]]
                temp_ind_without_A = [arr_test_ind_without_A[0]]

                for i in range(1, len(arr_test)):
                    if arr_test[i] < arr_test[i - 1]:
                        result_arr.append(temp)
                        result_arr_ind.append(temp_ind)
                        result_arr_ind_without_A.append(temp_ind_without_A)
                        temp = []
                        temp_ind = []
                        temp_ind_without_A = []
                    temp.append(arr_test[i])
                    temp_ind.append(arr_test_ind[i])
                    temp_ind_without_A.append(arr_test_ind_without_A[i])
                result_arr.append(temp)
                result_arr_ind.append(temp_ind)
                result_arr_ind_without_A.append(temp_ind_without_A)
                result_arr_ind_without_A_for_use = list(result_arr_ind_without_A)

                for i in range(len(result_arr_ind_without_A)):
                    result_arr_ind_without_A[i] = list(
                        range(result_arr_ind_without_A[i][0] - 1, result_arr_ind_without_A[i][-1] + 1))
                    result_arr_ind_without_A[i].append(result_arr_ind_without_A[i][-1] + 1)

                ########################################################################################################

                data_items = data.items()
                for key, value in data_items:
                    cell_group = value
                    column_group = str(cell_group)[0]
                    row_group = int(str(cell_group)[1:])

                    # ищем значения в ячейках ниже на одну или две от известной ячейки
                    # i - индекс количества подмассивов в массиве
                    data_from_data_rows1 = {}
                    data_from_data_rows2 = {}
                    data_from_data_aud1 = {}
                    data_from_data_aud2 = {}
                    for i in range(len(result_arr_ind_without_A)):
                        if row_group in result_arr_ind_without_A[i]:
                            data_from_rows1 = []
                            data_number_aud1 = []
                            data_from_rows2 = []
                            data_number_aud2 = []
                            for row in range(row_group + 1, max(result_arr_ind_without_A[i]) + 1):
                                if row in result_arr_ind_without_A_for_use[i]:
                                    finder = result_arr_ind_without_A_for_use[i].index(row)
                                    cell_info_A = 'A' + str(result_arr_ind_without_A_for_use[i][finder])
                                    cell_info = str(ws.cell(row=row, column=alphabet.find(column_group) + 1).value)
                                    cell_info2 = str(ws.cell(row=row, column=alphabet.find(column_group) + 3).value)
                                    cell_info2_1 = str(
                                        ws.cell(row=int(row) + 1, column=alphabet.find(column_group) + 3).value)

                                    data_from_rows1.append(str(cell_info))

                                    aud1_cell = str(result_arr_ind_without_A_for_use[i][finder])
                                    aud11_cell = str(
                                        ws.cell(row=int(aud1_cell), column=alphabet.find(column_group) + 2).value)
                                    aud12_cell = str(
                                        ws.cell(row=int(aud1_cell), column=alphabet.find(column_group) + 4).value)
                                    if cell_info2 != 'None':
                                        data_from_rows2.append(cell_info2)
                                    elif cell_info2 == 'None' and aud12_cell != 'None':
                                        data_from_rows2.append(cell_info)
                                    elif cell_info2 == 'None' and aud12_cell == 'None':
                                        data_from_rows2.append(cell_info2)
                                    elif cell_info2 != 'None' and cell_info2_1 != 'None':
                                        print("XYXYXXYXYY")
                                        data_from_rows2.append(cell_info2_1)

                                    if cell_info != 'None':
                                        # Заполняем массив значениями для первой подгруппы
                                        if aud11_cell != 'None':
                                            data_number_aud1.append(aud11_cell)
                                        elif aud11_cell == 'None':
                                            data_number_aud1.append(aud12_cell)
                                        else:
                                            data_number_aud1.append('None')
                                    else:
                                        data_number_aud1.append('None')

                                    if cell_info2 != 'None':
                                        if aud12_cell != 'None':
                                            data_number_aud2.append(aud12_cell)
                                        elif aud12_cell == 'None':
                                            data_number_aud2.append('None')
                                    elif cell_info2 == 'None' and cell_info != 'None':
                                        data_number_aud2.append(aud12_cell)
                                    else:
                                        data_number_aud2.append('None')
                                else:
                                    data_from_rows1.append(str(
                                        ws.cell(row=row, column=alphabet.find(column_group) + 1).value))
                                    cell_info = str(ws.cell(row=row, column=alphabet.find(column_group) + 1).value)
                                    cell_info2 = str(ws.cell(row=row, column=alphabet.find(column_group) + 3).value)
                                    cell_info2_1 = str(
                                        ws.cell(row=row, column=alphabet.find(column_group) + 3).value)
                                    if cell_info2 != 'None':
                                        data_from_rows2.append(str(cell_info2_1))
                                    else:
                                        data_from_rows2.append(str(cell_info))
                            data_from_data_rows1[key] = data_from_rows1
                            data_from_data_rows2[key] = data_from_rows2
                            data_from_data_aud1 = data_number_aud1
                            data_from_data_aud2 = data_number_aud2

                            all_data_collect[str(date_str)][key] = {}
                            all_data_collect[str(date_str)][key]['1'] = {}
                            all_data_collect[str(date_str)][key]['2'] = {}
                            all_data_collect[str(date_str)][key]['1']['lesson'] = data_from_data_rows1[key][0::2]
                            all_data_collect[str(date_str)][key]['1']['teacher'] = data_from_data_rows1[key][1::2]
                            all_data_collect[str(date_str)][key]['1']['aud'] = data_from_data_aud1
                            all_data_collect[str(date_str)][key]['2']['lesson'] = data_from_data_rows2[key][0::2]
                            all_data_collect[str(date_str)][key]['2']['teacher'] = data_from_data_rows2[key][1::2]
                            all_data_collect[str(date_str)][key]['2']['aud'] = data_from_data_aud2


                end = time.time()
                print("Время парсинга ", str(filename), ": ", end - start, "s")
        print("__________________________________")
    with open('all_data.json', 'w') as f:
        json.dump(all_data_collect, f)
    end_reader = time.time()
    print("Время полного выполнения Reader.py: ", end_reader - start_reader, "s")
