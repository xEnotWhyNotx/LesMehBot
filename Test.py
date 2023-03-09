import os
import openpyxl as ox
import numpy as np


data_dir = 'Admin/Downloads'
txt_files = [f for f in os.listdir(data_dir) if f.endswith('.txt')]
xlsx_files = [v for v in os.listdir(data_dir) if v.endswith('.xlsx')]
alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
dest_dir = 'Admin/Destination'
fds = sorted(os.listdir('Admin/Destination/'))
for file in fds:
    if file.endswith('.txt'):
        os.remove(dest_dir + '/' + file)

for file in txt_files:
    data = {}
    with open(os.path.join(data_dir, file), 'r') as f:
        lines = f.readlines()
        for line in lines:
            line = line.strip().split(': ')
            data[line[0]] = line[1]
    print(file)
    print(data)
    print("__________________________________")

    for filename in xlsx_files:
        if filename.endswith('.xlsx') and (str(filename)[:-4] + "txt") == file:
            #print(filename)
            filepath = os.path.join(data_dir, filename)
            wb = ox.load_workbook(filename=filepath, read_only=True)
            ws = wb[wb.sheetnames[0]]
            row_max = ws.max_row
            column_max = ws.max_column
            print('In file: ', filename, '\n Columns:', column_max, '\n Rows: ', row_max)
            row_min = 1
            column_min = 1

########################################################################################################################

            arr_A = []
            arr_indexes_A = []
            arr_indexes_without_A = []
            for i in range(row_max + 1):
                word_cell_A = 'A' + str(int(i))
                word_cell_without_A = int(i)
                arr_indexes_A.append(word_cell_A)
                arr_indexes_without_A.append(word_cell_without_A)
                # print(word_cell_A)
                data_from_cell_A = ws[word_cell_A].value
                arr_A.append(data_from_cell_A)

            NoneType = type(None)
            index_to_remove = []
            for j in range(len(arr_A)):
                if type(arr_A[j]) == NoneType:
                    index_to_remove.append(j)

            arr_A = np.delete(arr_A, index_to_remove)
            arr_indexes_A = np.delete(arr_indexes_A, index_to_remove)
            arr_indexes_without_A = np.delete(arr_indexes_without_A, index_to_remove)

            # Очистим массивы индексов и данных из ячеек от мусора
            next_clean_arr = []
            for i in range(len(arr_A)):
                nums = '123456'
                if arr_A[i] not in nums:
                    next_clean_arr.append(i)
            arr_A = np.delete(arr_A, next_clean_arr)
            arr_indexes_A = np.delete(arr_indexes_A, next_clean_arr)
            arr_indexes_without_A = np.delete(arr_indexes_without_A, next_clean_arr)
            # Функция создания массивов индексов и данных из ячеек
            arr_test = arr_A
            arr_test_ind = arr_indexes_A
            arr_test_ind_without_A = arr_indexes_without_A

            result_arr = []
            result_arr_ind = []
            result_arr_ind_without_A = []

            temp = [arr_test[0]]
            temp_ind = [arr_test_ind[0]]
            temp_ind_without_A = [arr_test_ind_without_A[0]]

            for i in range(1, len(arr_test)):
                if arr_test[i] < arr_test[i - 1]:
                    result_arr.append(temp)
                    result_arr_ind.append(temp_ind)
                    result_arr_ind_without_A.append(temp_ind_without_A)
                    temp = []
                    temp_ind = []
                    temp_ind_without_A = []
                temp.append(arr_test[i])
                temp_ind.append(arr_test_ind[i])
                temp_ind_without_A.append(arr_test_ind_without_A[i])
            result_arr.append(temp)
            result_arr_ind.append(temp_ind)
            result_arr_ind_without_A.append(temp_ind_without_A)
            result_arr_ind_without_A_for_use = list(result_arr_ind_without_A)

            for i in range(len(result_arr_ind_without_A)):
                result_arr_ind_without_A[i] = list(range(result_arr_ind_without_A[i][0] - 1, result_arr_ind_without_A[i][-1] + 1))
                result_arr_ind_without_A[i].append(result_arr_ind_without_A[i][-1] + 1)

            # print(arr_A)
            # print(arr_indexes_A)
            # print(arr_indexes_without_A)
            # print(result_arr)
            # print(result_arr_ind)
            #print(result_arr_ind_without_A)
            #print(result_arr_ind_without_A_for_use)

########################################################################################################################

            data_items = data.items()
            for key, value in data_items:
                cell_group = value
                column_group = str(cell_group)[0]
                row_group = int(str(cell_group)[1:])
                # print(row_group)
                # print(key)

                # ищем значения в ячейках ниже на одну или две от известной ячейки
                # i - индекс количества подмассивов в массиве
                # j - столбец в таблице
                data_from_data_rows = {}
                for i in range(len(result_arr_ind_without_A)):
                    if row_group in result_arr_ind_without_A[i]:
                        data_from_rows = []
                        #print(row_group, result_arr_ind_without_A[i])
                        for row in range(row_group + 1, max(result_arr_ind_without_A[i]) + 1):
                            if row in result_arr_ind_without_A_for_use[i]:
                                finder = result_arr_ind_without_A_for_use[i].index(row)
                                cell_info = 'A' + str(result_arr_ind_without_A_for_use[i][finder])
                                data_from_rows.append(str(ws[cell_info].value) + " " + str(ws.cell(row=row, column=alphabet.find(column_group) + 1).value))
                            else:
                                data_from_rows.append(ws.cell(row=row, column=alphabet.find(column_group) + 1).value)
                        # print(data_from_rows)
                        data_from_data_rows[key] = data_from_rows
                #print(data_from_data_rows)

                # Создаем текстовый файл с названием текущего xlsx файла
                with open(os.path.join(dest_dir, os.path.splitext(file)[0] + '_' + f'{key}' + '.txt'), 'w') as f:
                    for group, data in data_from_data_rows.items():
                        f.write(f"{group}: {data}\n")

                #print(result_arr_ind_without_A_for_use)
